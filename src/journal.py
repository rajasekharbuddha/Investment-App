"""
journal.py
==========
Appends new signals to the trading journal on OneDrive.
Runs automatically at the end of every daily scan.
"""

from pathlib import Path
from datetime import date, datetime
import os, sys
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False

from config import JOURNAL

JOURNAL_PATH  = Path(JOURNAL["PRIMARY_PATH"])
FALLBACK_PATH = Path(__file__).parent.parent / "reports" / JOURNAL["FALLBACK_NAME"]
SHEET_NAME    = JOURNAL["SHEET_SIGNALS"]
DATA_START_ROW = JOURNAL["DATA_START_ROW"]

FONT_NAME = "Arial"
INPUT_BG  = "FFFBEB"

def _inp(cell, value, fmt=None):
    cell.value = value
    cell.font  = Font(name=FONT_NAME, size=9, color="0000FF")
    cell.fill  = PatternFill("solid", start_color=INPUT_BG)
    cell.alignment = Alignment(wrap_text=True, vertical="center")
    if fmt:
        cell.number_format = fmt

def _fml(cell, formula, fmt=None):
    cell.value = formula
    cell.font  = Font(name=FONT_NAME, size=9, color="1F2937")
    cell.alignment = Alignment(vertical="center")
    if fmt:
        cell.number_format = fmt

# Column indices (1-based)
C_TRADENUM = 1;  C_EDATE  = 2;  C_ETIME  = 3;  C_XDATE  = 4
C_TICKER   = 5;  C_MARKET = 6;  C_DIR    = 7;  C_SETUP  = 8
C_CCY      = 9;  C_ENTRY  = 10; C_STOP   = 11; C_TARGET = 12
C_SHARES   = 13; C_EXIT   = 14; C_PNL    = 15; C_RMULT  = 16
C_SECTOR   = 17; C_REGIME = 18; C_ATRPCT = 19; C_GATES  = 20
C_CONF     = 21; C_SCREEN = 22; C_EMOTION= 23; C_CONFID = 24
C_PLAN     = 25; C_MISTAKE= 26; C_COND   = 27; C_QUALITY= 28
C_REFLECT  = 29


def _setup_name(regime: str) -> str:
    return {
        "Low Vol":  "Mastermind Pro — Low Vol",
        "Normal":   "Mastermind Pro — Normal",
        "High Vol": "Mastermind Pro — High Vol",
    }.get(regime, "Mastermind Pro Long")


def _resolve_path():
    if JOURNAL_PATH.exists():
        return JOURNAL_PATH
    if FALLBACK_PATH.exists():
        print(f"  [!] OneDrive path not found, using fallback: {FALLBACK_PATH}")
        return FALLBACK_PATH
    return None


def _reflection(d: dict, regime: str, atr_pct: float) -> str:
    parts = [f"Signal: {d['decision']}. {d.get('reason', '')}"]
    qs = d.get("quality_score")
    if qs is not None:
        from stock_selector import grade
        parts.append(f"Quality score: {qs:.0f} ({grade(qs)})")
    if d["decision"] == "NEAR":
        parts.append(f"NEAR: {d.get('gates_passed','?')}/5 gates passed. Monitor for setup completion.")
    if d["decision"] == "WAIT":
        parts.append("TODO: Complete §7.5 fundamentals check before upgrading to ENTER.")
    if d["decision"] == "ENTER":
        parts.append("TODO: Verify earnings >5 days, macro calendar, bid-ask spread before ordering.")
    if regime == "High Vol":
        parts.append("High Vol regime: ATR×3 stop, 2% risk — position smaller than normal.")
    if atr_pct and atr_pct > 3.5:
        parts.append(f"WARNING: ATR% {atr_pct:.2f} approaching Extreme (>4%). Monitor closely.")
    return " | ".join(parts)


def update_journal(decisions: list, watchlist: dict, markets: dict) -> str:
    if not OPENPYXL_OK:
        return "  [!] openpyxl not installed — run: python -m pip install openpyxl"

    journal_path = _resolve_path()
    if journal_path is None:
        return (
            f"  [!] Journal not found.\n"
            f"      Expected: {JOURNAL_PATH}\n"
            f"      Copy your journal there or update JOURNAL_PATH in src/config.py"
        )

    # Log ENTER, WAIT, and NEAR signals
    signals = [d for d in decisions if d["decision"] in ("ENTER", "WAIT", "NEAR")]
    if not signals:
        return "  [–] No ENTER/WAIT/NEAR signals today — nothing to log."

    today     = date.today()
    today_str = today.isoformat()

    try:
        wb = load_workbook(journal_path)
    except PermissionError:
        return "  [!] Journal is open in Excel — close it, then re-run."
    except Exception as e:
        return f"  [!] Could not open journal: {e}"

    if SHEET_NAME not in wb.sheetnames:
        return f"  [!] Sheet '{SHEET_NAME}' not found in journal."

    ws = wb[SHEET_NAME]

    first_empty = DATA_START_ROW
    for row_idx in range(DATA_START_ROW, DATA_START_ROW + 300):
        val = ws.cell(row=row_idx, column=C_TICKER).value
        if val is None or str(val).strip() == "":
            first_empty = row_idx
            break
    else:
        return "  [!] Trade log is full (300 rows). Add more rows to the journal."

    already_today = set()
    for row_idx in range(DATA_START_ROW, first_empty):
        d_val = ws.cell(row=row_idx, column=C_EDATE).value
        t_val = ws.cell(row=row_idx, column=C_TICKER).value
        d_date = d_val.date() if hasattr(d_val, "date") else d_val
        if t_val and str(d_date) == today_str:
            already_today.add(str(t_val))

    logged  = []
    skipped = []
    r = first_empty

    for d in signals:
        ticker = d["ticker"]
        if ticker in already_today:
            skipped.append(ticker)
            continue

        # Normalise watchlist lookup
        if isinstance(next(iter(watchlist.values()), None), list):
            from config import get_sector, get_market
            info   = {"market": get_market(ticker), "sector": get_sector(ticker)}
            market = info["market"]
            sector = info["sector"]
        else:
            info   = watchlist.get(ticker, {})
            market = info.get("market", "US")
            sector = info.get("sector", "")

        m_info   = markets.get(market, {})
        currency = m_info.get("currency", "USD")

        vol_gate = d.get("gates", {}).get("gate3_volatility",
                    d.get("gates", {}).get("volatility", {}))
        regime_d = vol_gate.get("regime", {}) if isinstance(vol_gate.get("regime"), dict) else {}
        regime   = regime_d.get("label", "") or vol_gate.get("details", {}).get("regime", "")

        atr_pct  = float(d.get("atr_pct", 0) or 0)
        price    = d.get("price", "")
        sizing   = d.get("sizing", {}) or {}
        stop     = sizing.get("stop", sizing.get("stop_price", ""))
        target   = sizing.get("target_2r", "")
        shares   = sizing.get("shares", "")

        gates_passed = d.get("gates_passed", sum(1 for g, res in d.get("gates", {}).items() if res.get("pass") is True))
        gates_total  = len(d.get("gates", {}))

        _fml(ws.cell(row=r, column=C_TRADENUM),
             f'=IF(E{r}<>"",IFERROR(MAX($A$5:A{r-1})+1,1),"")')
        _inp(ws.cell(row=r, column=C_EDATE),  today,   "yyyy-mm-dd")
        _inp(ws.cell(row=r, column=C_ETIME),  datetime.now().strftime("%H:%M"))
        _inp(ws.cell(row=r, column=C_XDATE),  "")
        _inp(ws.cell(row=r, column=C_TICKER), ticker)
        _inp(ws.cell(row=r, column=C_MARKET), market)
        _inp(ws.cell(row=r, column=C_DIR),    d["decision"])
        _inp(ws.cell(row=r, column=C_SETUP),  _setup_name(regime))
        _inp(ws.cell(row=r, column=C_CCY),    currency)
        _inp(ws.cell(row=r, column=C_ENTRY),  price,   "#,##0.00")
        _inp(ws.cell(row=r, column=C_STOP),   stop,    "#,##0.00")
        _inp(ws.cell(row=r, column=C_TARGET), target,  "#,##0.00")
        _inp(ws.cell(row=r, column=C_SHARES), shares,  "#,##0.0")
        _inp(ws.cell(row=r, column=C_EXIT),   "")

        _fml(ws.cell(row=r, column=C_PNL),
             f'=IF(OR(J{r}="",N{r}=""),"",(N{r}-J{r})*M{r}*IF(G{r}="Short",-1,1))',
             '#,##0.00;[Red](#,##0.00);"-"')
        _fml(ws.cell(row=r, column=C_RMULT),
             f'=IF(OR(J{r}="",N{r}="",K{r}="",J{r}=K{r}),"",'
             f'(N{r}-J{r})/ABS(J{r}-K{r})*IF(G{r}="Short",-1,1))',
             "0.00")

        _inp(ws.cell(row=r, column=C_SECTOR),  sector)
        _inp(ws.cell(row=r, column=C_REGIME),  regime)
        _inp(ws.cell(row=r, column=C_ATRPCT),  round(atr_pct, 2), "0.00")
        _inp(ws.cell(row=r, column=C_GATES),   f"{gates_passed}/{gates_total}")
        _inp(ws.cell(row=r, column=C_CONF),    str(gates_passed))
        _inp(ws.cell(row=r, column=C_SCREEN),  "")
        _inp(ws.cell(row=r, column=C_EMOTION), "")
        _inp(ws.cell(row=r, column=C_CONFID),  "")
        _inp(ws.cell(row=r, column=C_PLAN),    "N/A")
        _inp(ws.cell(row=r, column=C_MISTAKE), "")
        _inp(ws.cell(row=r, column=C_COND),    "")
        _inp(ws.cell(row=r, column=C_QUALITY), "")
        _inp(ws.cell(row=r, column=C_REFLECT), _reflection(d, regime, atr_pct))

        ws.row_dimensions[r].height = 60
        r += 1
        logged.append(f"{ticker} ({d['decision']})")

    try:
        wb.save(journal_path)
    except PermissionError:
        return "  [!] Save failed — close the journal in Excel first, then re-run."
    except Exception as e:
        return f"  [!] Save failed: {e}"

    lines = []
    if logged:
        lines.append(f"  [OK] Journal updated: {journal_path.name}")
        lines.append(f"      Logged {len(logged)} signal(s): {', '.join(logged)}")
        lines.append(f"      OneDrive will sync automatically in ~30 seconds")
    if skipped:
        lines.append(f"      Skipped (already logged today): {', '.join(skipped)}")
    return "\n".join(lines) if lines else "  [–] Nothing new to log."
