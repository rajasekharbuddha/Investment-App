"""
post_trade.py
=============
Enriches today's WAIT/ENTER/NEAR journal rows with:
  - Full position sizing (stop, target, shares) where missing
  - Market-behaviour analysis
  - Quality score context
  - Tier 3 reflection fields

Usage:
  python src/post_trade.py
"""

import sys
from pathlib import Path
from datetime import date

import pandas as pd

sys.path.insert(0, str(Path(__file__).parent))

from config import WATCHLIST_FLAT as WATCHLIST, MARKETS, ACCOUNT_SIZE_EUR
from data import fetch_history
from indicators import calculate_all, get_regime
from rules import calculate_position
from journal import (
    JOURNAL_PATH, FALLBACK_PATH, SHEET_NAME, DATA_START_ROW,
    C_EDATE, C_TICKER, C_STOP, C_TARGET, C_SHARES,
    C_EXIT, C_PNL, C_RMULT,
    C_EMOTION, C_CONFID, C_PLAN, C_MISTAKE,
    C_COND, C_QUALITY, C_REFLECT,
    _inp, _fml, _resolve_path,
)

try:
    from openpyxl import load_workbook
except ImportError:
    print("openpyxl not installed -- run: python -m pip install openpyxl")
    sys.exit(1)


def _analyse(df: pd.DataFrame) -> dict:
    row   = df.iloc[-1]
    prev1 = df.iloc[-2]
    prev5 = df.iloc[-6:-1]
    prev20 = df.iloc[-21:-1]

    close      = float(row["Close"])
    sma50      = float(row["SMA_50"])
    sma200     = float(row["SMA_200"])
    atr        = float(row["ATR"])
    atr_pct    = float(row["ATR_PCT"])
    rsi        = float(row["RSI"])
    macd_hist  = float(row["MACD_HIST"])
    volume     = float(row["Volume"])
    vol_avg20  = float(row["VOL_AVG_20"])

    dist_sma50  = (close - sma50)  / sma50  * 100
    dist_sma200 = (close - sma200) / sma200 * 100

    rsi_5ago  = float(prev5.iloc[0]["RSI"])
    rsi_trend = "rising" if rsi > rsi_5ago + 2 else ("falling" if rsi < rsi_5ago - 2 else "flat")

    hist_prev   = float(prev1["MACD_HIST"])
    macd_trend  = "improving" if macd_hist > hist_prev else "deteriorating"

    atr_5ago   = float(prev5.iloc[0]["ATR_PCT"])
    vol_regime = ("expanding" if atr_pct > atr_5ago * 1.05
                  else "contracting" if atr_pct < atr_5ago * 0.95
                  else "stable")

    price_5d  = (close - float(prev5.iloc[0]["Close"])) / float(prev5.iloc[0]["Close"]) * 100
    price_20d = (close - float(prev20.iloc[0]["Close"])) / float(prev20.iloc[0]["Close"]) * 100

    vol_ratio = volume / vol_avg20 if vol_avg20 > 0 else 1.0

    high_20   = float(prev20["High"].max())
    low_20    = float(prev20["Low"].min())
    pct_hi20  = (close - high_20) / high_20 * 100
    pct_lo20  = (close - low_20)  / low_20  * 100

    return dict(
        close=round(close, 2), sma50=round(sma50, 2), sma200=round(sma200, 2),
        atr=round(atr, 4), atr_pct=round(atr_pct, 2),
        dist_sma50=round(dist_sma50, 1), dist_sma200=round(dist_sma200, 1),
        rsi=round(rsi, 1), rsi_trend=rsi_trend,
        macd_hist=round(macd_hist, 4), macd_trend=macd_trend,
        vol_regime=vol_regime,
        price_5d=round(price_5d, 2), price_20d=round(price_20d, 2),
        vol_ratio=round(vol_ratio, 2),
        pct_hi20=round(pct_hi20, 1), pct_lo20=round(pct_lo20, 1),
        support=round(low_20, 2), resistance=round(high_20, 2),
    )


def _confidence(ctx: dict, regime: dict) -> int:
    score = 3
    if ctx["rsi_trend"]  == "rising":      score += 1
    if ctx["macd_trend"] == "improving":   score += 1
    if ctx["vol_regime"] == "contracting": score += 1
    if ctx["vol_ratio"]  > 1.2:            score -= 1
    if ctx["atr_pct"]    > 3.5:            score -= 1
    if ctx["price_5d"]   < -3:             score -= 1
    return max(1, min(5, score))


def _condition_tag(ctx: dict) -> str:
    d = ctx["dist_sma200"]
    if d >  5:  return "Uptrend — comfortably above SMA200"
    if d >  0:  return "Recovering — just reclaimed SMA200"
    if d > -5:  return "Testing SMA200 from below"
    return "Below SMA200 — early recovery watch"


def _quality_tag(score: int) -> str:
    return {5: "A+ Setup", 4: "A Setup", 3: "B Setup",
            2: "C Setup — marginal", 1: "D — avoid"}.get(score, "B Setup")


def _tier3_reflect(ticker: str, ctx: dict, regime: dict,
                   sizing: dict, score: int, market: str) -> str:
    info   = WATCHLIST.get(ticker, {})
    sector = info.get("sector", "")
    sym    = MARKETS.get(market, {}).get("symbol", "$")
    ccy    = MARKETS.get(market, {}).get("currency", "USD")
    lines  = []

    lines.append(
        f"PRICE: {sym}{ctx['close']:,.2f}  |  "
        f"SMA50 {sym}{ctx['sma50']:,.2f} ({ctx['dist_sma50']:+.1f}%)  "
        f"SMA200 {sym}{ctx['sma200']:,.2f} ({ctx['dist_sma200']:+.1f}%)"
    )
    lines.append(
        f"MOMENTUM: RSI {ctx['rsi']} ({ctx['rsi_trend']})  |  "
        f"MACD histogram {ctx['macd_trend']}  |  "
        f"5-day {ctx['price_5d']:+.1f}%  20-day {ctx['price_20d']:+.1f}%"
    )
    lines.append(
        f"VOLATILITY: ATR {ctx['atr_pct']:.2f}% ({ctx['vol_regime']})  |  "
        f"Regime: {regime['regime']}  |  "
        f"Volume {ctx['vol_ratio']:.1f}x 20-day avg"
    )
    lines.append(
        f"LEVELS: 20-day support {sym}{ctx['support']:,.2f}  "
        f"({ctx['pct_lo20']:+.1f}% from close)  |  "
        f"20-day resistance {sym}{ctx['resistance']:,.2f}  "
        f"({ctx['pct_hi20']:+.1f}% from close)"
    )
    if sizing.get("can_trade"):
        lines.append(
            f"SIZING ({ccy}): Entry {sym}{sizing['entry']:,.2f}  "
            f"Stop {sym}{sizing['stop']:,.2f} ({regime['stop_mult']}xATR)  "
            f"Target 2R {sym}{sizing['target_2r']:,.2f}  "
            f"Shares {sizing['shares']:.1f}  "
            f"Position {sym}{sizing['position_val']:,.0f}  "
            f"Risk {regime['risk_pct']}% = {sym}{sizing['risk_amount']:,.0f}"
        )
        proj_pnl = sizing["risk_amount"] * 2
        lines.append(
            f"PROJECTED (2R hit): Exit {sym}{sizing['target_2r']:,.2f}  "
            f"P&L +{sym}{proj_pnl:,.0f}  R-Multiple +2.0R"
        )
        lines.append(
            f"DOWNSIDE (stop hit): Exit {sym}{sizing['stop']:,.2f}  "
            f"P&L -{sym}{sizing['risk_amount']:,.0f}  R-Multiple -1.0R"
        )

    if ctx["dist_sma200"] >= 0:
        plan = ("All automated gates PASS. Verify §7.5 fundamentals "
                "(revenue YoY>0, EPS YoY>0, FCF>0, D/E<2, no earnings <5 days). "
                "If confirmed: place limit order at or below today's close.")
    else:
        plan = ("Price still below SMA200. WAIT — do not enter until price "
                "reclaims SMA200 AND fundamentals pass §7.5.")
    lines.append(f"ACTION: {plan}")

    if ctx["atr_pct"] > 3.5:
        lines.append(
            f"RISK FLAG: ATR% {ctx['atr_pct']:.2f} is close to Extreme threshold (4.0%). "
            "Position is already halved by regime rule. Monitor daily."
        )

    return "\n".join(lines)


def run():
    journal_path = _resolve_path()
    if journal_path is None:
        print(f"[!] Journal not found.\n    Expected: {JOURNAL_PATH}")
        return

    try:
        wb = load_workbook(journal_path)
    except PermissionError:
        print("[!] Journal is open in Excel -- close it first.")
        return
    except Exception as e:
        print(f"[!] Could not open journal: {e}")
        return

    if SHEET_NAME not in wb.sheetnames:
        print(f"[!] Sheet '{SHEET_NAME}' not found.")
        return

    ws = wb[SHEET_NAME]
    today_str = date.today().isoformat()

    rows = []
    for r in range(DATA_START_ROW, DATA_START_ROW + 300):
        d_val = ws.cell(row=r, column=C_EDATE).value
        t_val = ws.cell(row=r, column=C_TICKER).value
        if t_val is None or str(t_val).strip() == "":
            break
        d_date = d_val.date() if hasattr(d_val, "date") else d_val
        if str(d_date) == today_str:
            rows.append((r, str(t_val).strip()))

    if not rows:
        print(f"No rows for today ({today_str}) found in journal.")
        return

    print(f"\nFound {len(rows)} signal(s) for {today_str}: "
          f"{', '.join(t for _, t in rows)}\n")
    print("=" * 68)

    updated = []
    for row_idx, ticker in rows:
        print(f"\n  {ticker}")
        print("  " + "-" * 40)

        try:
            df = calculate_all(fetch_history(ticker, years=3))
        except Exception as e:
            print(f"  [!] Data error: {e}")
            continue

        ctx    = _analyse(df)
        regime = get_regime(ctx["atr_pct"])
        sizing = calculate_position(ACCOUNT_SIZE_EUR, ctx["close"], ctx["atr"], regime)
        info   = WATCHLIST.get(ticker, {})
        market = info.get("market", "US")
        sym    = MARKETS.get(market, {}).get("symbol", "$")
        score  = _confidence(ctx, regime)

        print(f"  Regime:       {regime['regime']}  ATR%={ctx['atr_pct']:.2f}%  ({ctx['vol_regime']})")
        print(f"  RSI:          {ctx['rsi']} ({ctx['rsi_trend']})   MACD: {ctx['macd_trend']}")
        print(f"  5d / 20d:     {ctx['price_5d']:+.1f}% / {ctx['price_20d']:+.1f}%")
        print(f"  vs SMA50:     {ctx['dist_sma50']:+.1f}%   vs SMA200: {ctx['dist_sma200']:+.1f}%")
        print(f"  Volume:       {ctx['vol_ratio']:.1f}x avg")
        if sizing.get("can_trade"):
            print(f"  Entry:        {sym}{sizing['entry']:,.2f}")
            print(f"  Stop:         {sym}{sizing['stop']:,.2f}  ({regime['stop_mult']}x ATR = {sym}{ctx['atr']:,.2f})")
            print(f"  Target 2R:    {sym}{sizing['target_2r']:,.2f}")
            print(f"  Shares:       {sizing['shares']:.1f}")
            print(f"  Position:     {sym}{sizing['position_val']:,.0f}  "
                  f"(risk {regime['risk_pct']}% = {sym}{sizing['risk_amount']:,.0f})")
            proj_pnl = sizing["risk_amount"] * 2
            print(f"  Proj P&L 2R:  +{sym}{proj_pnl:,.0f}  (+2.0R)")
        print(f"  Confidence:   {score}/5  ({_quality_tag(score)})")
        print(f"  Condition:    {_condition_tag(ctx)}")

        if sizing.get("can_trade"):
            if not ws.cell(row=row_idx, column=C_STOP).value:
                _inp(ws.cell(row=row_idx, column=C_STOP),   sizing["stop"],      "#,##0.00")
                _inp(ws.cell(row=row_idx, column=C_TARGET), sizing["target_2r"], "#,##0.00")
                _inp(ws.cell(row=row_idx, column=C_SHARES), sizing["shares"],    "#,##0.0")

        _inp(ws.cell(row=row_idx, column=C_EMOTION), "Disciplined — protocol followed, no chasing")
        _inp(ws.cell(row=row_idx, column=C_CONFID),  f"{score}/5")
        _inp(ws.cell(row=row_idx, column=C_PLAN),
             "WAIT: complete §7.5 fundamentals check. "
             "ENTER: place limit at or below today's close if all pass. "
             "SKIP: remove from watchlist if any fundamental fails.")
        _inp(ws.cell(row=row_idx, column=C_MISTAKE), "")
        _inp(ws.cell(row=row_idx, column=C_COND),    _condition_tag(ctx))
        _inp(ws.cell(row=row_idx, column=C_QUALITY), _quality_tag(score))
        _inp(ws.cell(row=row_idx, column=C_REFLECT),
             _tier3_reflect(ticker, ctx, regime, sizing, score, market))

        ws.row_dimensions[row_idx].height = 100
        updated.append(ticker)

    try:
        wb.save(journal_path)
        print(f"\n{'='*68}")
        print(f"[OK] Journal updated: {', '.join(updated)}")
        print(f"     File: {journal_path.name}")
        print(f"     OneDrive will sync automatically in ~30 seconds.")
    except PermissionError:
        print("\n[!] Save failed -- close the journal in Excel first.")
    except Exception as e:
        print(f"\n[!] Save failed: {e}")


if __name__ == "__main__":
    run()
